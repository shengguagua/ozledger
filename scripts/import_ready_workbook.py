#!/usr/bin/env python3
import argparse
import json
import re
import sys
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def parse_settings(sheet):
    settings = {}
    for key, value in sheet.iter_rows(min_row=2, values_only=True):
        if key:
            settings[str(key)] = value
    return {
        "exchangeRate": float(settings.get("exchangeRate") or 4.5),
        "usdRate": float(settings.get("usdRate") or 1.5),
        "financialNote": str(settings.get("financialNote") or ""),
    }


def read_accounts(sheet):
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values[0]:
            continue
        row = dict(zip(headers, values))
        rows.append(
            {
                "id": str(row["id"]),
                "name": str(row["name"]),
                "owner": str(row["owner"]),
                "type": str(row["type"]),
                "currency": str(row["currency"]),
                "initialBalance": float(row["balance"] or 0),
                "iconName": str(row["iconName"] or "Wallet"),
                "color": str(row["color"] or "") or None,
            }
        )
    return rows


def derive_date_rates(unused_sheet, default_rate):
    rates = {}
    for row in unused_sheet.iter_rows(min_row=2, values_only=True):
        date_hint = row[2]
        label = str(row[4] or "")
        raw_value = str(row[5] or "")
        if not date_hint:
            continue
        date_hint = str(date_hint)
        if "澳元汇率" in label:
            try:
                rates[date_hint] = float(raw_value)
                continue
            except ValueError:
                pass

        combined = f"{label} {raw_value}"
        match = re.search(r"1\\s*AUD\\s*=\\s*([0-9.]+)", combined)
        if match:
            rates[date_hint] = float(match.group(1))
            continue

        match = re.search(r"人民币([0-9.]+)", combined)
        if match:
            rates[date_hint] = float(match.group(1))

    return defaultdict(lambda: default_rate, rates)


def build_snapshots(sheet, accounts, exchange_rates, usd_rate):
    headers = [cell.value for cell in sheet[1]]
    account_map = {account["id"]: account for account in accounts}
    grouped = defaultdict(list)

    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values[0]:
            continue
        row = dict(zip(headers, values))
        grouped[str(row["date"])].append(row)

    snapshots = []
    for date, rows in sorted(grouped.items()):
        merged_rows = {}
        for row in rows:
            key = str(row["accountId"])
            if key not in merged_rows:
                merged_rows[key] = {
                    "date": date,
                    "accountId": key,
                    "balance": float(row["balance"] or 0),
                    "note": str(row["note"]) if row.get("note") else "",
                }
            else:
                merged_rows[key]["balance"] += float(row["balance"] or 0)
                if row.get("note"):
                    extra_note = str(row["note"])
                    if extra_note not in merged_rows[key]["note"]:
                        merged_rows[key]["note"] = " | ".join(filter(None, [merged_rows[key]["note"], extra_note]))

        details = []
        total_cny = 0.0
        note = ""
        for row in merged_rows.values():
            account = account_map[row["accountId"]]
            balance = float(row["balance"] or 0)
            if row.get("note") and not note:
                note = str(row["note"])

            details.append(
                {
                    "accountId": account["id"],
                    "name": account["name"],
                    "owner": account["owner"],
                    "type": account["type"],
                    "balance": balance,
                    "currency": account["currency"],
                }
            )

            if account["currency"] == "CNY":
                cny_value = balance
            elif account["currency"] == "AUD":
                cny_value = balance * exchange_rates[date]
            elif account["currency"] == "USD":
                cny_value = balance * usd_rate * exchange_rates[date]
            else:
                cny_value = balance

            if account["type"] in {"credit", "huabei"}:
                total_cny -= cny_value
            else:
                total_cny += cny_value

        snapshots.append(
            {
                "id": f"snap-{date}",
                "date": date,
                "totalCNY": round(total_cny, 2),
                "note": note,
                "isDeleted": False,
                "accountDetails": details,
            }
        )

    snapshots.sort(key=lambda item: item["date"], reverse=True)
    return snapshots


def build_payload(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    required = {"accounts", "snapshot_rows", "settings", "unused_data", "issues"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise ValueError(f"Workbook missing required sheets: {', '.join(sorted(missing))}")

    settings = parse_settings(wb["settings"])
    accounts = read_accounts(wb["accounts"])
    exchange_rates = derive_date_rates(wb["unused_data"], settings["exchangeRate"])
    snapshots = build_snapshots(wb["snapshot_rows"], accounts, exchange_rates, settings["usdRate"])

    return {
      "accounts": accounts,
      "snapshots": snapshots,
      "transactions": [],
      "exchangeRate": settings["exchangeRate"],
      "usdRate": settings["usdRate"],
      "financialNote": settings["financialNote"],
    }


def post_payload(api_url, payload):
    request = urllib.request.Request(
        api_url,
        method="POST",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(request) as response:
        return json.loads(response.read().decode("utf-8"))


def main():
    parser = argparse.ArgumentParser(description="Convert an OZLedger-ready workbook into import payload JSON and optionally import it.")
    parser.add_argument("workbook", help="Path to the cleaned xlsx workbook")
    parser.add_argument("--json-out", help="Write normalized import JSON to this path")
    parser.add_argument("--api-url", help="POST the normalized import payload to this API endpoint, e.g. http://127.0.0.1:8787/api/backup/import")
    args = parser.parse_args()

    payload = build_payload(args.workbook)

    if args.json_out:
        out_path = Path(args.json_out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    if args.api_url:
        result = post_payload(args.api_url, payload)
        print(json.dumps(result.get("meta", {}), ensure_ascii=False, indent=2))
    else:
        print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Import failed: {exc}", file=sys.stderr)
        sys.exit(1)
