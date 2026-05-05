#!/usr/bin/env python3
import argparse
import hashlib
import json
import sys
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


COMMON_IDS = {
    ("小盛", "工行卡", "CNY", "bank"): "xs_icbc_cny",
    ("小盛", "招行卡", "CNY", "bank"): "xs_cmb_cny",
    ("小盛", "支付宝", "CNY", "alipay"): "xs_alipay_cny",
    ("小盛", "微信", "CNY", "wechat"): "xs_wechat_cny",
    ("小盛", "现金", "CNY", "cash"): "xs_cash_cny",
    ("小盛", "现金", "AUD", "cash"): "xs_cash_aud",
    ("小盛", "花呗", "CNY", "huabei"): "xs_huabei_cny",
    ("小盛", "京东白条", "CNY", "credit"): "xs_jdbaitiao_cny",
    ("小盛", "广发白金卡", "CNY", "credit"): "xs_gf_credit_cny",
    ("小盛", "银行卡（工，中）-工", "CNY", "bank"): "xs_icbc_cny",
    ("小盛", "银行卡（工，中）-中", "CNY", "bank"): "xs_boc_cn_cny",
    ("小盛", "银行卡-主", "CNY", "bank"): "xs_bankcard_primary_cny",
    ("小盛", "银行卡-副", "CNY", "bank"): "xs_bankcard_secondary_cny",
    ("小盛", "工行", "AUD", "bank"): "xs_icbc_aud",
    ("小盛", "中行", "AUD", "bank"): "xs_boc_cn_aud",
    ("小盛", "BOC", "AUD", "bank"): "xs_boc_aud",
    ("小盛", "HSBC", "AUD", "bank"): "xs_hsbc_aud",
    ("小盛", "ANZ", "AUD", "bank"): "xs_anz_aud",
    ("小盛", "COM", "AUD", "bank"): "xs_commbank_aud",
    ("大王", "工行卡", "CNY", "bank"): "dw_icbc_cny",
    ("大王", "建行卡", "CNY", "bank"): "dw_ccb_cny",
    ("大王", "支付宝", "CNY", "alipay"): "dw_alipay_cny",
    ("大王", "微信", "CNY", "wechat"): "dw_wechat_cny",
    ("大王", "现金", "CNY", "cash"): "dw_cash_cny",
    ("大王", "现金", "AUD", "cash"): "dw_cash_aud",
    ("大王", "花呗", "CNY", "huabei"): "dw_huabei_cny",
    ("大王", "白条", "CNY", "credit"): "dw_baitiao_cny",
    ("大王", "银行卡（工，中）-工", "CNY", "bank"): "dw_icbc_cny",
    ("大王", "银行卡（工，中）-中", "CNY", "bank"): "dw_boc_cn_cny",
    ("大王", "银行卡-主", "CNY", "bank"): "dw_bankcard_primary_cny",
    ("大王", "银行卡-副", "CNY", "bank"): "dw_bankcard_secondary_cny",
    ("大王", "工行", "AUD", "bank"): "dw_icbc_aud",
    ("大王", "中行", "AUD", "bank"): "dw_boc_cn_aud",
    ("大王", "BOC", "AUD", "bank"): "dw_boc_aud",
    ("大王", "HSBC", "AUD", "bank"): "dw_hsbc_aud",
    ("大王", "ANZ", "AUD", "bank"): "dw_anz_aud",
    ("大王", "COM", "AUD", "bank"): "dw_commbank_aud",
    ("家庭", "A股票", "CNY", "investment"): "sp_a_stock_cny",
    ("家庭", "美股", "AUD", "investment"): "sp_us_stock_aud",
    ("家庭", "bond", "AUD", "longterm"): "sp_bond_aud",
    ("家庭", "房租押金", "AUD", "longterm"): "sp_rent_bond_aud",
    ("家庭", "预付房租", "AUD", "longterm"): "sp_prepaid_rent_aud",
}

KEEP_SPECIAL_NAMES = {
    "A股票",
    "美股",
    "bond",
    "房租押金",
    "预付房租",
    "外汇",
    "外汇（印尼卢比）",
    "银行卡",
    "信用卡",
}

COLLAPSE_KEYWORDS = (
    "改口费",
    "彩礼",
    "礼金",
    "券",
    "美团",
    "点评",
    "抖音",
    "旅游",
    "机酒",
    "留学",
    "婚庆",
    "结婚",
    "欠款",
    "出借",
    "iphone",
    "cpu",
    "翻译件",
    "申请押金",
    "手机",
    "电脑",
    "学费",
    "生活费",
    "酒店",
    "房租",
    "sony",
    "夏",
    "刘洋",
    "彭亮",
    "盛正干",
    "淮上楼",
    "京东金融",
    "白条",
    "白金卡",
    "大王",
    "盛伟存款",
    "妈存钱",
)


def owner_prefix(owner: str) -> str:
    return {"小盛": "xs", "大王": "dw", "家庭": "sp"}.get(owner, "sp")


def stable_id(owner: str, name: str, currency: str, account_type: str) -> str:
    key = (owner, name, currency, account_type)
    if key in COMMON_IDS:
      return COMMON_IDS[key]
    digest = hashlib.md5("|".join(key).encode("utf-8")).hexdigest()[:10]
    return f"{owner_prefix(owner)}_{digest}"


@dataclass
class SnapshotBuilder:
    accounts: dict = field(default_factory=dict)
    latest_dates: dict = field(default_factory=dict)
    snapshots: list = field(default_factory=list)

    def add_detail(self, details, date, owner, name, balance, currency, account_type):
        if balance is None:
            return
        balance = float(balance)
        account_id = stable_id(owner, name, currency, account_type)
        details.append({
            "accountId": account_id,
            "name": name,
            "owner": owner,
            "type": account_type,
            "balance": balance,
            "currency": currency,
        })

        if account_id not in self.latest_dates or date >= self.latest_dates[account_id]:
            self.latest_dates[account_id] = date
            self.accounts[account_id] = {
                "id": account_id,
                "name": name,
                "owner": owner,
                "type": account_type,
                "currency": currency,
                "initialBalance": balance,
                "iconName": "Wallet",
                "color": None,
            }

    def add_snapshot(self, date, total_cny, details, note=""):
        merged = {}
        for detail in details:
            key = detail["accountId"]
            if key not in merged:
                merged[key] = detail.copy()
            else:
                merged[key]["balance"] += detail["balance"]
        self.snapshots.append({
            "id": f"snap-{date}",
            "date": date,
            "totalCNY": round(float(total_cny), 2),
            "note": note,
            "isDeleted": False,
            "accountDetails": list(merged.values()),
        })


def v(ws, row, col):
    return ws.cell(row=row, column=col).value


def count_snapshot_presence(snapshots):
    presence = {}
    for snapshot in snapshots:
        seen = set()
        for detail in snapshot["accountDetails"]:
            seen.add(detail["accountId"])
        for account_id in seen:
            presence[account_id] = presence.get(account_id, 0) + 1
    return presence


def should_collapse(detail, presence_count):
    name = detail["name"]
    account_type = detail["type"]

    if any(keyword in name for keyword in COLLAPSE_KEYWORDS):
        return True
    if name in KEEP_SPECIAL_NAMES:
        return False
    if account_type in {"bank", "alipay", "wechat", "cash", "huabei"}:
        return False
    if account_type == "investment" and name in {"A股票", "美股"}:
        return False
    if account_type == "longterm" and name in {"bond", "房租押金", "预付房租"}:
        return False
    if presence_count >= 2 and account_type in {"credit", "investment", "longterm"}:
        return False
    return presence_count <= 1


def format_note_chunk(detail):
    balance = round(float(detail["balance"]), 2)
    currency = detail["currency"]
    owner = detail["owner"]
    return f"{owner}-{detail['name']} {balance:g} {currency}"


def collapse_snapshot_details(payload):
    presence = count_snapshot_presence(payload["snapshots"])
    kept_accounts_meta = {}

    for snapshot in payload["snapshots"]:
        kept_details = []
        collapsed_chunks = []
        for detail in snapshot["accountDetails"]:
            if should_collapse(detail, presence.get(detail["accountId"], 0)):
                collapsed_chunks.append(format_note_chunk(detail))
                continue
            kept_details.append(detail)
            account_id = detail["accountId"]
            if account_id not in kept_accounts_meta:
                kept_accounts_meta[account_id] = {
                    "id": account_id,
                    "name": detail["name"],
                    "owner": detail["owner"],
                    "type": detail["type"],
                    "currency": detail["currency"],
                    "iconName": "Wallet",
                    "color": None,
                }

        snapshot["accountDetails"] = kept_details
        if collapsed_chunks:
            collapsed_note = "已折叠一次性项目: " + "；".join(collapsed_chunks)
            snapshot["note"] = f"{snapshot['note']} | {collapsed_note}".strip(" |")

    latest_detail_map = {}
    if payload["snapshots"]:
        latest_detail_map = {
            detail["accountId"]: float(detail["balance"])
            for detail in payload["snapshots"][0]["accountDetails"]
        }

    payload["accounts"] = sorted(
        [
            {
                **account,
                "initialBalance": latest_detail_map.get(account_id, 0.0),
            }
            for account_id, account in kept_accounts_meta.items()
        ],
        key=lambda item: (item["owner"], item["name"], item["currency"]),
    )
    return payload


def parse_modern_sheet(ws, date, total_cell, builder):
    details = []
    # CNY core
    builder.add_detail(details, date, "大王", "银行卡（工，中）-工", v(ws, 2, 2), "CNY", "bank")
    if v(ws, 2, 3) is not None:
        builder.add_detail(details, date, "大王", "银行卡（工，中）-中", v(ws, 2, 3), "CNY", "bank")
    builder.add_detail(details, date, "大王", "信用卡", v(ws, 2, 4), "CNY", "credit")
    # detect extra fx column on 2025.05.20
    if ws.title == "2025.05.20余额":
        builder.add_detail(details, date, "家庭", "外汇（印尼卢比）", v(ws, 2, 5), "CNY", "cash")
        builder.add_detail(details, date, "大王", "微信", v(ws, 2, 6), "CNY", "wechat")
        builder.add_detail(details, date, "大王", "支付宝", v(ws, 2, 7), "CNY", "alipay")
        builder.add_detail(details, date, "大王", "花呗", v(ws, 2, 8), "CNY", "huabei")
        builder.add_detail(details, date, "大王", "现金", v(ws, 2, 9), "CNY", "cash")

        builder.add_detail(details, date, "小盛", "银行卡（工，中）-工", v(ws, 3, 2), "CNY", "bank")
        builder.add_detail(details, date, "小盛", "银行卡（工，中）-中", v(ws, 3, 3), "CNY", "bank")
        builder.add_detail(details, date, "小盛", "信用卡", v(ws, 3, 4), "CNY", "credit")
        builder.add_detail(details, date, "家庭", "外汇（印尼卢比）", v(ws, 3, 5), "CNY", "cash")
        builder.add_detail(details, date, "小盛", "微信", v(ws, 3, 6), "CNY", "wechat")
        builder.add_detail(details, date, "小盛", "支付宝", v(ws, 3, 7), "CNY", "alipay")
        builder.add_detail(details, date, "小盛", "花呗", v(ws, 3, 8), "CNY", "huabei")
        builder.add_detail(details, date, "小盛", "现金", v(ws, 3, 9), "CNY", "cash")
    elif ws.title in {"2025.1.24余额", "2025.03.25余额", "2025.04.04余额", "2025.04.17余额"}:
        builder.add_detail(details, date, "大王", "微信", v(ws, 2, 6), "CNY", "wechat")
        builder.add_detail(details, date, "大王", "支付宝", v(ws, 2, 7), "CNY", "alipay")
        builder.add_detail(details, date, "大王", "花呗", v(ws, 2, 8), "CNY", "huabei")
        builder.add_detail(details, date, "大王", "现金", v(ws, 2, 9), "CNY", "cash")

        builder.add_detail(details, date, "小盛", "银行卡（工，中）-工", v(ws, 3, 2), "CNY", "bank")
        if v(ws, 3, 3) is not None:
            builder.add_detail(details, date, "小盛", "银行卡（工，中）-中", v(ws, 3, 3), "CNY", "bank")
        builder.add_detail(details, date, "小盛", "信用卡", v(ws, 3, 4), "CNY", "credit")
        builder.add_detail(details, date, "小盛", "微信", v(ws, 3, 6), "CNY", "wechat")
        builder.add_detail(details, date, "小盛", "支付宝", v(ws, 3, 7), "CNY", "alipay")
        builder.add_detail(details, date, "小盛", "花呗", v(ws, 3, 8), "CNY", "huabei")
        builder.add_detail(details, date, "小盛", "现金", v(ws, 3, 9), "CNY", "cash")
    else:
        builder.add_detail(details, date, "大王", "微信", v(ws, 2, 5), "CNY", "wechat")
        builder.add_detail(details, date, "大王", "支付宝", v(ws, 2, 6), "CNY", "alipay")
        builder.add_detail(details, date, "大王", "花呗", v(ws, 2, 7), "CNY", "huabei")
        builder.add_detail(details, date, "大王", "现金", v(ws, 2, 8), "CNY", "cash")

        builder.add_detail(details, date, "小盛", "银行卡（工，中）-工", v(ws, 3, 2), "CNY", "bank")
        if v(ws, 3, 3) is not None:
            builder.add_detail(details, date, "小盛", "银行卡（工，中）-中", v(ws, 3, 3), "CNY", "bank")
        builder.add_detail(details, date, "小盛", "信用卡", v(ws, 3, 4), "CNY", "credit")
        builder.add_detail(details, date, "小盛", "微信", v(ws, 3, 5), "CNY", "wechat")
        builder.add_detail(details, date, "小盛", "支付宝", v(ws, 3, 6), "CNY", "alipay")
        builder.add_detail(details, date, "小盛", "花呗", v(ws, 3, 7), "CNY", "huabei")
        builder.add_detail(details, date, "小盛", "现金", v(ws, 3, 8), "CNY", "cash")

    # AUD core
    aud_cols = [(2, "工行"), (3, "中行"), (4, "BOC"), (5, "HSBC"), (6, "ANZ"), (7, "COM"), (8, "现金")]
    for col, name in aud_cols:
        builder.add_detail(details, date, "小盛", name, v(ws, 6, col), "AUD", "cash" if name == "现金" else "bank")
        builder.add_detail(details, date, "大王", name, v(ws, 7, col), "AUD", "cash" if name == "现金" else "bank")

    # Special rows that contribute to totals
    if ws.title in {"2025.1.24余额", "2025.03.25余额", "2025.04.04余额", "2025.04.17余额", "2025.05.20余额", "2025.07.10余额", "2025.09.19余额"}:
        if v(ws, 10, 2) not in (None, 0):
            builder.add_detail(details, date, "家庭", str(v(ws, 10, 1)), v(ws, 10, 2), "CNY", "investment")
        if v(ws, 10, 3) not in (None, 0):
            builder.add_detail(details, date, "家庭", str(v(ws, 10, 1)), v(ws, 10, 3), "AUD", "investment")
        if v(ws, 11, 2) not in (None, 0):
            builder.add_detail(details, date, "家庭", str(v(ws, 11, 1)), v(ws, 11, 2), "CNY", "investment")
        if v(ws, 11, 3) not in (None, 0):
            builder.add_detail(details, date, "家庭", str(v(ws, 11, 1)), v(ws, 11, 3), "AUD", "investment")

    # sheet-specific included extras
    if ws.title == "2025.1.24余额":
        builder.add_detail(details, date, "家庭", "学费（人民币）", v(ws, 14, 2), "CNY", "credit")
        builder.add_detail(details, date, "家庭", "学费（澳元）", v(ws, 14, 3), "AUD", "credit")
        builder.add_detail(details, date, "家庭", "2.15-2.28房租", v(ws, 15, 3), "AUD", "credit")
        builder.add_detail(details, date, "家庭", "未报销", v(ws, 18, 2), "CNY", "pending")
        builder.add_detail(details, date, "家庭", "房租押金", v(ws, 19, 3), "AUD", "longterm")
    elif ws.title == "2025.03.25余额":
        builder.add_detail(details, date, "家庭", "学费（人民币）", v(ws, 14, 2), "CNY", "credit")
        builder.add_detail(details, date, "家庭", "学费（澳元）", v(ws, 14, 3), "AUD", "credit")
        builder.add_detail(details, date, "家庭", "妈存钱", v(ws, 15, 4), "AUD", "credit")
        builder.add_detail(details, date, "家庭", "bond", v(ws, 18, 3), "AUD", "longterm")
    elif ws.title == "2025.04.04余额":
        builder.add_detail(details, date, "家庭", "妈存钱", v(ws, 15, 4), "AUD", "credit")
        builder.add_detail(details, date, "家庭", "bond", v(ws, 18, 3), "AUD", "longterm")
    elif ws.title == "2025.04.17余额":
        builder.add_detail(details, date, "家庭", "妈存钱", v(ws, 15, 4), "AUD", "credit")
        builder.add_detail(details, date, "家庭", "bond", v(ws, 18, 3), "AUD", "longterm")
    elif ws.title == "2025.05.20余额":
        for row in (14, 15, 16):
            if v(ws, row, 2) not in (None, 0):
                builder.add_detail(details, date, "家庭", str(v(ws, row, 1)), v(ws, row, 2), "CNY", "pending")
            if v(ws, row, 3) not in (None, 0):
                builder.add_detail(details, date, "家庭", str(v(ws, row, 1)), v(ws, row, 3), "AUD", "pending")
        builder.add_detail(details, date, "家庭", "bond", v(ws, 31, 3), "AUD", "longterm")
    elif ws.title == "2025.07.10余额":
        builder.add_detail(details, date, "家庭", "sony 2470 gm2", v(ws, 14, 2), "CNY", "pending")
        for row in (19, 20, 21, 22):
            builder.add_detail(details, date, "家庭", str(v(ws, row, 1)), v(ws, row, 2), "CNY", "pending")
        builder.add_detail(details, date, "家庭", "bond", v(ws, 30, 3), "AUD", "longterm")
    elif ws.title == "2025.09.19余额":
        builder.add_detail(details, date, "家庭", "携程酒店", v(ws, 14, 2), "CNY", "pending")
        builder.add_detail(details, date, "家庭", "9月生活费", v(ws, 19, 2), "CNY", "pending")
        builder.add_detail(details, date, "家庭", "bond", v(ws, 27, 3), "AUD", "longterm")

    total = v(ws, total_cell[0], total_cell[1])
    builder.add_snapshot(date, total, details)


def parse_2024_10_13(ws, builder):
    date = "2024-10-13"
    details = []
    for owner, row in [("大王", 2), ("小盛", 3)]:
        for col, name in [(2, "HSBC"), (3, "BOC"), (4, "ANZ"), (5, "现金")]:
            builder.add_detail(details, date, owner, name, v(ws, row, col), "AUD", "cash" if name == "现金" else "bank")
    for row in (8, 9, 10, 11):
        builder.add_detail(details, date, "家庭", str(v(ws, row, 1)), v(ws, row, 4), "AUD", "investment" if str(v(ws, row, 1)) == "美股" else "pending")
    builder.add_detail(details, date, "大王", "微信", v(ws, 18, 2), "CNY", "wechat")
    builder.add_detail(details, date, "大王", "支付宝", v(ws, 18, 3), "CNY", "alipay")
    builder.add_detail(details, date, "大王", "A股票", v(ws, 18, 4), "CNY", "investment")
    builder.add_detail(details, date, "大王", "银行卡", v(ws, 18, 5), "CNY", "bank")
    builder.add_detail(details, date, "小盛", "微信", v(ws, 19, 2), "CNY", "wechat")
    builder.add_detail(details, date, "小盛", "支付宝", v(ws, 19, 3), "CNY", "alipay")
    builder.add_detail(details, date, "小盛", "A股票", v(ws, 19, 4), "CNY", "investment")
    builder.add_detail(details, date, "小盛", "银行卡", v(ws, 19, 5), "CNY", "bank")
    builder.add_snapshot(date, v(ws, 2, 8), details)


def parse_2024_10_02(ws, builder):
    date = "2024-10-02"
    details = []
    for owner, row in [("大王", 2), ("小盛", 3)]:
        for col, name, typ in [(2, "HSBC", "bank"), (3, "BOC", "bank"), (4, "投资", "investment"), (5, "BOND", "longterm"), (6, "现金", "cash")]:
            builder.add_detail(details, date, owner, name, v(ws, row, col), "AUD", typ)
        builder.add_detail(details, date, owner, "微信", v(ws, row, 10), "CNY", "wechat")
        builder.add_detail(details, date, owner, "支付宝", v(ws, row, 11), "CNY", "alipay")
        builder.add_detail(details, date, owner, "银行卡", v(ws, row, 12), "CNY", "bank")
    builder.add_snapshot(date, v(ws, 2, 15), details)


def parse_2024_09_23(ws, builder):
    date = "2024-09-23"
    details = []
    for owner, row in [("大王", 2), ("小盛", 3)]:
        builder.add_detail(details, date, owner, "HSBC", v(ws, row, 2), "AUD", "bank")
        builder.add_detail(details, date, owner, "BOC", v(ws, row, 3), "AUD", "bank")
        builder.add_detail(details, date, owner, "微信", v(ws, row, 7), "CNY", "wechat")
        builder.add_detail(details, date, owner, "支付宝", v(ws, row, 8), "CNY", "alipay")
        builder.add_detail(details, date, owner, "银行卡", v(ws, row, 9), "CNY", "bank")
    builder.add_snapshot(date, v(ws, 2, 12), details)


def parse_2024_08_01(ws, builder):
    date = "2024-08-01"
    details = []
    builder.add_detail(details, date, "大王", "HSBC", v(ws, 2, 2), "AUD", "bank")
    builder.add_detail(details, date, "大王", "BOC", v(ws, 2, 3), "AUD", "bank")
    builder.add_detail(details, date, "大王", "支付宝", v(ws, 2, 5), "CNY", "alipay")
    builder.add_detail(details, date, "大王", "微信", v(ws, 2, 6), "CNY", "wechat")
    builder.add_detail(details, date, "大王", "中行", v(ws, 2, 7), "CNY", "bank")
    builder.add_detail(details, date, "小盛", "HSBC", v(ws, 3, 2), "AUD", "bank")
    builder.add_detail(details, date, "小盛", "BOC", v(ws, 3, 3), "AUD", "bank")
    builder.add_detail(details, date, "小盛", "现金", v(ws, 3, 4), "AUD", "cash")
    builder.add_detail(details, date, "小盛", "待报销", v(ws, 3, 8), "CNY", "pending")
    builder.add_snapshot(date, v(ws, 4, 1), details)


def parse_2023_10_26(ws, builder):
    date = "2023-10-26"
    details = []
    builder.add_detail(details, date, "大王", "支付宝", v(ws, 2, 2), "CNY", "alipay")
    builder.add_detail(details, date, "大王", "微信", (v(ws, 3, 2) or 0) + (v(ws, 4, 2) or 0), "CNY", "wechat")
    builder.add_detail(details, date, "大王", "工行卡", v(ws, 5, 2), "CNY", "bank")
    builder.add_detail(details, date, "大王", "建行卡", v(ws, 6, 2), "CNY", "bank")
    builder.add_detail(details, date, "大王", "抖音券", v(ws, 7, 2), "CNY", "pending")
    builder.add_detail(details, date, "大王", "美团", v(ws, 8, 2), "CNY", "pending")
    builder.add_detail(details, date, "大王", "大众点评", v(ws, 9, 2), "CNY", "pending")
    builder.add_detail(details, date, "大王", "花呗", v(ws, 13, 2), "CNY", "huabei")
    builder.add_detail(details, date, "大王", "白条", v(ws, 14, 2), "CNY", "credit")
    builder.add_detail(details, date, "大王", "盛伟存款", v(ws, 15, 2), "CNY", "credit")
    builder.add_detail(details, date, "大王", "彩礼", v(ws, 16, 2), "CNY", "credit")
    builder.add_detail(details, date, "大王", "盛改口费", v(ws, 17, 2), "CNY", "credit")
    builder.add_detail(details, date, "大王", "王改口费", v(ws, 18, 2), "CNY", "credit")
    builder.add_detail(details, date, "大王", "礼金", v(ws, 19, 2), "CNY", "credit")
    total = (
        float(v(ws, 2, 2) or 0)
        + float(v(ws, 3, 2) or 0)
        + float(v(ws, 4, 2) or 0)
        + float(v(ws, 5, 2) or 0)
        + float(v(ws, 6, 2) or 0)
        + float(v(ws, 7, 2) or 0)
        + float(v(ws, 8, 2) or 0)
        - float(v(ws, 13, 2) or 0)
        - float(v(ws, 14, 2) or 0)
        - float(v(ws, 15, 2) or 0)
        - float(v(ws, 16, 2) or 0)
        - float(v(ws, 17, 2) or 0)
        - float(v(ws, 18, 2) or 0)
        - float(v(ws, 19, 2) or 0)
    )
    builder.add_snapshot(date, total, details, note="王敬荷个人快照")


def parse_2023_10_08(ws, builder):
    date = "2023-10-08"
    details = []
    for row, name, t in [
        (3, "工行卡", "bank"), (4, "招行卡", "bank"), (5, "支付宝", "alipay"), (6, "微信", "wechat"),
        (7, "京东金融", "investment"), (8, "旅游未报销（机酒）", "pending"), (9, "泰铢", "cash"),
        (10, "留学未报销", "pending"), (11, "婚庆未报销", "pending"), (12, "大王欠款", "pending"), (14, "券", "pending")
    ]:
        builder.add_detail(details, date, "小盛", name, v(ws, row, 2), "CNY", t)
    builder.add_snapshot(date, v(ws, 2, 2), details, note="小盛个人快照")


def parse_personal_pair(sheng_ws, wang_ws, date, sheng_total_row, wang_total_row, builder):
    details = []
    # Sheng
    for row, name, t in [
        (3, "工行卡", "bank"), (4, "招行卡", "bank"), (5, "支付宝", "alipay"), (6, "微信", "wechat"),
        (7, "现金", "cash"), (8, "出借（彭亮）", "longterm"), (9, "刘洋", "pending"), (10, "出借（盛正干）", "longterm"),
        (13, "iphone15promax256", "pending"), (14, "cpu", "pending"), (15, "美团", "pending"),
        (16, "淮上楼储蓄卡", "pending"), (17, "抖音", "pending"), (20, "翻译件", "pending"),
        (21, "学校申请押金", "pending"), (22, "结婚未报销", "pending"), (23, "王手机", "pending"), (24, "盛电脑", "pending")
    ]:
        val = v(sheng_ws, row, 2)
        if val is not None:
            builder.add_detail(details, date, "小盛", name, val, "CNY", t)
    for row, name, t in [
        (2, "花呗", "huabei"), (3, "京东白条", "credit"), (4, "广发白金卡", "credit"),
        (5, "大王", "credit"), (8, "盛改口费", "credit"), (9, "学费", "credit")
    ]:
        val = v(sheng_ws, row, 4)
        if val is not None:
            builder.add_detail(details, date, "小盛", name, val, "CNY", t)

    # Wang
    for row, name, t in [
        (3, "工行卡", "bank"), (4, "建行卡", "bank"), (5, "支付宝", "alipay"),
        (6, "盛欠款", "pending"), (7, "微信", "wechat"), (10, "美团", "pending"),
        (11, "大众点评", "pending"), (12, "抖音", "pending")
    ]:
        val = v(wang_ws, row, 2)
        if val is not None:
            builder.add_detail(details, date, "大王", name, val, "CNY", t)
    for row, name, t in [
        (2, "花呗", "huabei"), (8, "礼金改口费（王妈）", "credit"), (9, "彩礼", "credit"), (10, "PTE学费", "credit")
    ]:
        val = v(wang_ws, row, 4)
        if val is not None:
            builder.add_detail(details, date, "大王", name, val, "CNY", t)

    total = float(v(sheng_ws, sheng_total_row, 2) or 0) + float(v(wang_ws, wang_total_row, 2) or 0)
    builder.add_snapshot(date, total, details)


def parse_2024_04_19(ws, builder):
    date = "2024-04-19"
    details = []
    for row, name, t in [
        (3, "工行卡", "bank"), (4, "招行卡", "bank"), (5, "支付宝", "alipay"), (6, "微信", "wechat"),
        (7, "现金", "cash"), (10, "外汇", "cash"), (13, "美团", "pending"), (14, "淮上楼储蓄卡", "pending"),
        (15, "抖音", "pending"), (16, "夏", "pending")
    ]:
        builder.add_detail(details, date, "小盛", name, v(ws, row, 2), "CNY", t)
    for row, name, t in [
        (2, "花呗", "huabei"), (3, "京东白条", "credit"), (4, "广发白金卡", "credit"),
        (5, "大王", "credit")
    ]:
        builder.add_detail(details, date, "小盛", name, v(ws, row, 4), "CNY", t)
    builder.add_detail(details, date, "小盛", "盛改口费", v(ws, 2, 7), "CNY", "credit")
    builder.add_snapshot(date, v(ws, 22, 2), details, note="小盛个人快照")


def build_payload(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    builder = SnapshotBuilder()

    parse_2023_10_08(wb["2023年10月8日（盛伟）"], builder)
    parse_2023_10_26(wb["2023年10月26日（王敬荷）"], builder)
    parse_personal_pair(wb["2024年1月15日（盛伟）"], wb["2024年1月15日（王敬荷）"], "2024-01-15", 27, 15, builder)
    parse_personal_pair(wb["2024年1月18日（盛伟）"], wb["2024年1月18日（王敬荷）"], "2024-01-18", 23, 15, builder)
    parse_2024_04_19(wb["2024年4月19日（盛伟）"], builder)
    parse_2024_08_01(wb["2024.08.01余额"], builder)
    parse_2024_09_23(wb["2024.09.23余额"], builder)
    parse_2024_10_02(wb["2024.10.02余额"], builder)
    parse_2024_10_13(wb["2024.10.13余额"], builder)
    parse_modern_sheet(wb["2025.1.24余额"], "2025-01-24", (14, 11), builder)
    parse_modern_sheet(wb["2025.03.25余额"], "2025-03-25", (14, 11), builder)
    parse_modern_sheet(wb["2025.04.04余额"], "2025-04-04", (14, 11), builder)
    parse_modern_sheet(wb["2025.04.17余额"], "2025-04-17", (14, 11), builder)
    parse_modern_sheet(wb["2025.05.20余额"], "2025-05-20", (13, 11), builder)
    parse_modern_sheet(wb["2025.07.10余额"], "2025-07-10", (13, 11), builder)
    parse_modern_sheet(wb["2025.09.19余额"], "2025-09-19", (6, 12), builder)

    builder.snapshots.sort(key=lambda item: item["date"], reverse=True)
    accounts = sorted(builder.accounts.values(), key=lambda item: (item["owner"], item["name"], item["currency"]))
    payload = {
        "accounts": accounts,
        "snapshots": builder.snapshots,
        "transactions": [],
        "exchangeRate": 4.5,
        "usdRate": 1.0,
        "financialNote": "从原始账单.xlsx直接解析导入。2026.03.05 / 2026.01.27 / 2025.12.08 为图片页，待手工补录。",
    }
    return collapse_snapshot_details(payload)


def post_payload(api_url, payload):
    request = urllib.request.Request(
        api_url,
        method="POST",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(request) as response:
        return json.loads(response.read().decode("utf-8"))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--json-out")
    parser.add_argument("--api-url")
    args = parser.parse_args()
    payload = build_payload(args.workbook)
    if args.json_out:
        out = Path(args.json_out)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.api_url:
        print(json.dumps(post_payload(args.api_url, payload).get("meta", {}), ensure_ascii=False, indent=2))
    else:
        print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Import failed: {exc}", file=sys.stderr)
        raise
